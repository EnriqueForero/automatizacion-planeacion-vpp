# -*- coding: utf-8 -*-
"""
Constructores de Excel para las 3 familias.

Generan las matrices de captura vacías (o pre-llenadas)
que las áreas diligencian cada trimestre.

Clases:
    ConstructorExcelMisional     — 8 hojas (VP, Marca País)
    ConstructorExcelTerritorial  — 22 hojas (Hub, Oficom, OfiReg)
    ConstructorExcelTransversal  — ~27 hojas (Gerencias transversales)
"""

from __future__ import annotations

import os
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from .config import (
    Config, TipoUnidad, FamiliaUnidad,
    EjeMisional, EJES_REFERENCIA, ORDEN_EJES,
)
from .utils import EstilosExcel, _ss, _trunc, _safe_filename, log

# ═══════════════════════════════════════════════════════════════════════
# CONSTRUCTOR EXCEL — MISIONAL
# ═══════════════════════════════════════════════════════════════════════

class ConstructorExcelMisional:
    """
    Genera la matriz Excel de captura para unidades MISIONALES.
    Responsabilidad única: crear el archivo Excel con estructura correcta.
    """

    def __init__(self, config: Config):
        self.cfg = config
        self.E = EstilosExcel

    def generar(
        self,
        nombre_unidad: str,
        tipo_unidad: str,
        trimestre: str = 'Q1',
        anio: str = '2026',
        num_lineas: int = 5,
        datos_base: Optional[Dict] = None,
        output_dir: Optional[str] = None
    ) -> str:
        """
        Genera un Excel de captura para una unidad misional.

        Args:
            nombre_unidad: Nombre del área (ej. 'VP Exportaciones')
            tipo_unidad: EJE o MARCA PAÍS
            trimestre: Q1, Q2, Q3 o Q4
            anio: Año del seguimiento
            num_lineas: Líneas estratégicas a crear (1-5)
            datos_base: Dict con datos pre-llenados (opcional)
            output_dir: Carpeta de salida

        Returns:
            Ruta del archivo generado
        """
        num_lineas = max(1, min(num_lineas, self.cfg.max_lineas_estrategicas))
        output_dir = output_dir or self.cfg.dir_entrada
        wb = openpyxl.Workbook()

        self._crear_instrucciones(wb, nombre_unidad, num_lineas)
        self._crear_configuracion(wb, nombre_unidad, tipo_unidad, trimestre, anio)
        self._crear_dofa(wb, nombre_unidad, datos_base)
        self._crear_tendencias(wb, nombre_unidad, datos_base)

        for le in range(1, num_lineas + 1):
            self._crear_linea_estrategica(wb, nombre_unidad, le, datos_base)

        self._crear_casos_exito(wb, nombre_unidad, datos_base)
        self._crear_metas(wb, nombre_unidad, datos_base)
        self._crear_monitoreo(wb, nombre_unidad, num_lineas)

        # Guardar
        safe = _safe_filename(nombre_unidad)
        filename = f'{trimestre}_{anio}_{tipo_unidad}_{safe}.xlsx'
        path = os.path.join(output_dir, filename)
        wb.save(path)
        log.info(f"✅ Excel generado: {filename} ({num_lineas} líneas)")
        return path

    # --- Hojas individuales ---

    def _crear_instrucciones(self, wb, nombre: str, num_lineas: int):
        ws = wb.active
        ws.title = 'INSTRUCCIONES'
        ws.sheet_properties.tabColor = '1B3A5C'
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 70

        ws.merge_cells('B2:C2')
        ws['B2'].value = 'MATRIZ DE CAPTURA — SEGUIMIENTO TRIMESTRAL'
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color='1B3A5C')
        ws.merge_cells('B3:C3')
        ws['B3'].value = f'{nombre} — ProColombia'
        ws['B3'].font = Font(name='Calibri', size=12, color='666666')

        instrucciones = [
            ('¿Qué es esta matriz?',
             'Formato oficial para reportar avances trimestrales de la '
             'Hoja de Ruta. Reemplaza el diligenciamiento en PowerPoint. '
             'Complete las celdas azules.'),
            ('Hojas a diligenciar',
             '1) CONFIGURACIÓN → Verifique datos\n'
             '2) DOFA → Estado de cada ítem\n'
             '3) TENDENCIAS → Actualice si aplica\n'
             f'4) LÍNEAS ESTRATÉGICAS (1 a {num_lineas}) → Avances\n'
             '5) CASOS DE ÉXITO → Destacados del Q\n'
             '6) METAS → Avance acumulado'),
            ('Código de colores',
             '🔵 Azul claro = Usted diligencia\n'
             '🟡 Amarillo = Info base (no editar)\n'
             '⬜ Gris = Automático\n'
             '🔴 Rojo claro = ¡Pendiente!'),
            ('Reglas',
             f'• Máximo {self.cfg.max_chars_campo} caracteres por campo\n'
             '• Use listas desplegables para estados\n'
             '• No modifique la estructura\n'
             '• Si una línea no aplica, déjela vacía\n'
             '• MONITOREO muestra completitud en tiempo real'),
        ]
        for i, (t, d) in enumerate(instrucciones):
            r = 5 + i * 3
            ws.cell(row=r, column=2, value=t).font = self.E.FNT_H2
            ws.merge_cells(start_row=r, start_column=3,
                          end_row=r + 1, end_column=3)
            c = ws.cell(row=r, column=3, value=d)
            c.font = self.E.FNT_N
            c.alignment = self.E.ALN_W
        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_configuracion(self, wb, nombre, tipo, trimestre, anio):
        ws = wb.create_sheet('CONFIGURACIÓN')
        ws.sheet_properties.tabColor = '2E6DA4'
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40

        ws.merge_cells('A1:B1')
        ws['A1'].value = 'CONFIGURACIÓN GENERAL'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR

        campos = [
            ('Trimestre en seguimiento', trimestre),
            ('Año', anio),
            ('Tipo de unidad', tipo),
            ('Nombre de la unidad', nombre),
            ('Fecha de envío lineamientos', ''),
            ('Fecha límite diligenciamiento', ''),
        ]
        for i, (lbl, val) in enumerate(campos):
            r = 3 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=lbl),
                          font=self.E.FNT_LBL, fill=self.E.FL_LCK)
            self.E.aplicar(ws.cell(row=r, column=2, value=val),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)

        # Validaciones
        dv_q = DataValidation(type='list', formula1='"Q1,Q2,Q3,Q4"',
                              allow_blank=False)
        ws.add_data_validation(dv_q)
        dv_q.add('B3')

        tipos_str = ','.join(TipoUnidad.valores_validos())
        dv_t = DataValidation(type='list', formula1=f'"{tipos_str}"')
        ws.add_data_validation(dv_t)
        dv_t.add('B5')

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_dofa(self, wb, nombre: str, datos: Optional[Dict]):
        ws = wb.create_sheet('DOFA')
        ws.sheet_properties.tabColor = 'C0392B'
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 55

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'SEGUIMIENTO DOFA — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR
        ws.merge_cells('A2:D2')
        ws['A2'].value = ('Col B = Hoja de Ruta original. '
                          'Col C = Estado. Col D = Actualización si aplica.')
        ws['A2'].font = self.E.FNT_SM

        for c, h in enumerate(
            ['ID', 'Información Hoja de Ruta (Base)',
             'Estado Seguimiento', 'Actualización / Cambio'], 1
        ):
            self.E.aplicar(ws.cell(row=4, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        dv_d = DataValidation(
            type='list',
            formula1='"Se mantiene,Se elimina,Se actualiza"',
            allow_blank=True
        )
        ws.add_data_validation(dv_d)

        cuads = ['DEBILIDADES', 'OPORTUNIDADES', 'FORTALEZAS', 'AMENAZAS']
        fills_c = {'DEB': 'C0392B', 'OPO': '2E6DA4',
                   'FOR': '27AE60', 'AME': 'F39C12'}
        n = self.cfg.max_dofa_por_cuadrante

        r = 5
        for cq in cuads:
            pre = cq[:3]
            ws.merge_cells(start_row=r, start_column=1,
                          end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value=cq)
            c.font = self.E.FNT_WB
            c.fill = PatternFill('solid', fgColor=fills_c[pre])
            c.alignment = self.E.ALN_C
            r += 1

            # Pre-llenar con datos base si existen
            items_base = []
            if datos and 'dofa' in datos:
                items_base = datos['dofa'].get(cq, [])

            for i in range(n):
                self.E.aplicar(ws.cell(row=r, column=1, value=f'{pre}-{i+1}'),
                              font=self.E.FNT_SM, fill=self.E.FL_LCK)
                base_text = items_base[i] if i < len(items_base) else ''
                self.E.aplicar(ws.cell(row=r, column=2, value=base_text),
                              fill=self.E.FL_BAS)
                self.E.aplicar(ws.cell(row=r, column=3),
                              fill=self.E.FL_EDT, prot=self.E.PROT_U)
                dv_d.add(f'C{r}')
                self.E.aplicar(ws.cell(row=r, column=4),
                              fill=self.E.FL_EDT, prot=self.E.PROT_U)
                r += 1

        # Formato condicional
        for rx in range(6, r):
            ws.conditional_formatting.add(
                f'C{rx}',
                FormulaRule(
                    formula=[f'AND(NOT(ISBLANK(B{rx})),ISBLANK(C{rx}))'],
                    fill=self.E.FL_RED
                )
            )
            ws.conditional_formatting.add(
                f'D{rx}',
                FormulaRule(
                    formula=[f'AND(C{rx}="Se actualiza",ISBLANK(D{rx}))'],
                    fill=self.E.FL_RED
                )
            )
        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_tendencias(self, wb, nombre: str, datos: Optional[Dict]):
        ws = wb.create_sheet('TENDENCIAS')
        ws.sheet_properties.tabColor = '8E44AD'
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 65
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 65

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'SEGUIMIENTO TENDENCIAS — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR

        for c, h in enumerate(
            ['#', 'Tendencia (Base)', 'Estado', 'Actualización'], 1
        ):
            self.E.aplicar(ws.cell(row=3, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        dv_te = DataValidation(type='list',
                               formula1='"Se mantiene,Se actualiza"',
                               allow_blank=True)
        ws.add_data_validation(dv_te)

        items_base = (datos or {}).get('tendencias', [])

        for i in range(self.cfg.max_tendencias):
            r = 4 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_LBL, fill=self.E.FL_LCK)
            base_text = items_base[i] if i < len(items_base) else ''
            self.E.aplicar(ws.cell(row=r, column=2, value=base_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=3),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            dv_te.add(f'C{r}')
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_linea_estrategica(self, wb, nombre: str, le_num: int,
                                  datos: Optional[Dict]):
        sn = f'LÍNEA ESTRATÉGICA {le_num}'
        ws = wb.create_sheet(sn)
        idx = min(le_num - 1, len(self.E.TAB_COLORS) - 1)
        ws.sheet_properties.tabColor = self.E.TAB_COLORS[idx]

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 20

        ws.merge_cells('A1:E1')
        ws['A1'].value = f'{sn} — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR

        # Nombre de la línea
        ws['A2'].value = 'Nombre de la línea:'
        ws['A2'].font = self.E.FNT_LBL
        ws.merge_cells('B2:E2')

        le_data = None
        if datos and 'lineas' in datos and le_num <= len(datos['lineas']):
            le_data = datos['lineas'][le_num - 1]
            self.E.aplicar(ws['B2'], fill=self.E.FL_EDT,
                          prot=self.E.PROT_U, font=self.E.FNT_H2)
            ws['B2'].value = le_data.get('nombre', '')
        else:
            self.E.aplicar(ws['B2'], fill=self.E.FL_EDT,
                          prot=self.E.PROT_U, font=self.E.FNT_H2)

        # Sección acciones
        ws.merge_cells('A4:E4')
        ws['A4'].value = 'ACCIONES Y ACTIVIDADES'
        ws['A4'].font = self.E.FNT_WB
        ws['A4'].fill = self.E.FL_SUB

        headers_acc = ['#', 'Acción (Hoja de Ruta)',
                       'Actividades Clave (Hoja de Ruta)',
                       'Avance del Q (texto)', 'Estado']
        for c, h in enumerate(headers_acc, 1):
            self.E.aplicar(
                ws.cell(row=5, column=c, value=h),
                font=self.E.FNT_WB,
                fill=PatternFill('solid', fgColor='34495E'),
                align=self.E.ALN_C
            )

        dv_ea = DataValidation(
            type='list',
            formula1='"Completada,En progreso,Sin avance,No aplica"',
            allow_blank=True
        )
        ws.add_data_validation(dv_ea)

        acciones_base = le_data.get('acciones', []) if le_data else []
        max_acc = self.cfg.max_acciones_por_linea

        for i in range(max_acc):
            r = 6 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_SM, fill=self.E.FL_LCK)
            acc_text = acciones_base[i].get('accion', '') \
                if i < len(acciones_base) else ''
            act_text = acciones_base[i].get('actividad', '') \
                if i < len(acciones_base) else ''
            self.E.aplicar(ws.cell(row=r, column=2, value=acc_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=3, value=act_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=5),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U,
                          align=self.E.ALN_C)
            dv_ea.add(f'E{r}')

        # Formato condicional para avances
        for rx in range(6, 6 + max_acc):
            ws.conditional_formatting.add(
                f'D{rx}',
                FormulaRule(
                    formula=[f'AND(NOT(ISBLANK(B{rx})),ISBLANK(D{rx}))'],
                    fill=self.E.FL_RED
                )
            )
            ws.conditional_formatting.add(
                f'D{rx}',
                FormulaRule(
                    formula=[f'AND(NOT(ISBLANK(B{rx})),NOT(ISBLANK(D{rx})))'],
                    fill=self.E.FL_GRN
                )
            )

        # Sección indicadores
        ir = 6 + max_acc + 1
        ws.merge_cells(f'A{ir}:E{ir}')
        ws.cell(row=ir, column=1,
                value='INDICADORES DE ÉXITO / RESULTADOS')
        ws.cell(row=ir, column=1).font = self.E.FNT_WB
        ws.cell(row=ir, column=1).fill = self.E.FL_SUB

        ih = ir + 1
        for c, h in enumerate(
            ['#', 'Indicador', 'Meta', 'Avance Acumulado', 'Observaciones'], 1
        ):
            self.E.aplicar(
                ws.cell(row=ih, column=c, value=h),
                font=self.E.FNT_WB,
                fill=PatternFill('solid', fgColor='34495E'),
                align=self.E.ALN_C
            )

        inds_base = le_data.get('indicadores', []) if le_data else []
        max_ind = self.cfg.max_indicadores_por_linea

        for i in range(max_ind):
            r = ih + 1 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_SM, fill=self.E.FL_LCK)
            ind_text = inds_base[i].get('indicador', '') \
                if i < len(inds_base) else ''
            meta_text = inds_base[i].get('meta', '') \
                if i < len(inds_base) else ''
            self.E.aplicar(ws.cell(row=r, column=2, value=ind_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=3, value=meta_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=5),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_casos_exito(self, wb, nombre: str, datos: Optional[Dict]):
        ws = wb.create_sheet('CASOS DE ÉXITO')
        ws.sheet_properties.tabColor = 'F39C12'
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 75
        ws.column_dimensions['D'].width = 22

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'CASOS DE ÉXITO — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR

        for c, h in enumerate(
            ['#', 'Título del caso',
             'Descripción (máx 500 car.)', 'Eje/Línea relacionada'], 1
        ):
            self.E.aplicar(ws.cell(row=3, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        for i in range(self.cfg.max_casos_exito):
            r = 4 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_SM, fill=self.E.FL_LCK)
            self.E.aplicar(ws.cell(row=r, column=2),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=3),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U,
                          align=self.E.ALN_C)

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_metas(self, wb, nombre: str, datos: Optional[Dict]):
        ws = wb.create_sheet('METAS GENERALES')
        ws.sheet_properties.tabColor = '16A085'
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 15

        ws.merge_cells('A1:E1')
        ws['A1'].value = f'RESUMEN DE METAS GENERALES — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR

        for c, h in enumerate(
            ['#', 'Indicador', 'Meta Anual',
             'Avance Acumulado', '% Avance'], 1
        ):
            self.E.aplicar(ws.cell(row=3, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        metas_base = (datos or {}).get('metas', [])

        for i in range(self.cfg.max_metas):
            r = 4 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_SM, fill=self.E.FL_LCK)
            ind_text = metas_base[i].get('indicador', '') \
                if i < len(metas_base) else ''
            meta_text = metas_base[i].get('meta', '') \
                if i < len(metas_base) else ''
            self.E.aplicar(ws.cell(row=r, column=2, value=ind_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=3, value=meta_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            ws.cell(row=r, column=5).value = \
                f'=IF(AND(C{r}<>"",C{r}<>0,D{r}<>""),D{r}/C{r},"")'
            self.E.aplicar(ws.cell(row=r, column=5),
                          font=self.E.FNT_FRM, fill=self.E.FL_LCK)
            ws.cell(row=r, column=5).number_format = '0.0%'

        # Formato condicional
        lr = 4 + self.cfg.max_metas - 1
        ws.conditional_formatting.add(
            f'E4:E{lr}',
            CellIsRule(operator='greaterThan', formula=['0.8'],
                      fill=self.E.FL_GRN))
        ws.conditional_formatting.add(
            f'E4:E{lr}',
            CellIsRule(operator='between', formula=['0.5', '0.8'],
                      fill=PatternFill('solid', fgColor='FEF9E7')))
        ws.conditional_formatting.add(
            f'E4:E{lr}',
            CellIsRule(operator='lessThan', formula=['0.5'],
                      fill=self.E.FL_RED))

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_monitoreo(self, wb, nombre: str, num_lineas: int):
        ws = wb.create_sheet('MONITOREO')
        ws.sheet_properties.tabColor = '2C3E50'
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'PANEL DE MONITOREO — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR
        ws['A2'].value = 'Este panel se calcula automáticamente. No editar.'
        ws['A2'].font = self.E.FNT_SM

        for c, h in enumerate(
            ['Sección', 'Ítems con datos', 'Total esperado',
             '% Completitud'], 1
        ):
            self.E.aplicar(ws.cell(row=4, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        # Fila de referencia para filas DOFA:
        # Cuadrantes: cada uno tiene max_dofa filas
        # Fila 5=encabezado DEB, 6-15=items, 16=encab OPO, 17-26, etc.
        n = self.cfg.max_dofa_por_cuadrante  # 10
        # Las filas de items DOFA: 6:15, 17:26, 28:37, 39:48
        # Para C (estado): contar no blancos en C6:C{5+4*(n+1)-1}
        dofa_last = 5 + 4 * (n + 1) - 1  # =48 para n=10

        secs = [
            ('DOFA — Estados',
             f'=COUNTIF(DOFA!C6:C{dofa_last},"Se mantiene")'
             f'+COUNTIF(DOFA!C6:C{dofa_last},"Se elimina")'
             f'+COUNTIF(DOFA!C6:C{dofa_last},"Se actualiza")',
             f'=COUNTA(DOFA!B6:B{dofa_last})'
             f'-COUNTBLANK(DOFA!B6:B{dofa_last})'),
            ('Tendencias — Estados',
             '=COUNTA(TENDENCIAS!C4:C10)-COUNTBLANK(TENDENCIAS!C4:C10)',
             '=COUNTA(TENDENCIAS!B4:B10)-COUNTBLANK(TENDENCIAS!B4:B10)'),
        ]

        max_acc = self.cfg.max_acciones_por_linea
        acc_first = 6
        acc_last = 5 + max_acc  # =17 para 12

        max_ind = self.cfg.max_indicadores_por_linea
        ind_first = acc_last + 3  # header row + 1
        ind_last = ind_first + max_ind - 1

        for n_le in range(1, num_lineas + 1):
            sn = f"'LÍNEA ESTRATÉGICA {n_le}'"
            secs.append((
                f'Línea {n_le} — Avances',
                f'=COUNTA({sn}!D{acc_first}:D{acc_last})'
                f'-COUNTBLANK({sn}!D{acc_first}:D{acc_last})',
                f'=COUNTA({sn}!B{acc_first}:B{acc_last})'
                f'-COUNTBLANK({sn}!B{acc_first}:B{acc_last})'
            ))
            secs.append((
                f'Línea {n_le} — Indicadores',
                f'=COUNTA({sn}!D{ind_first}:D{ind_last})'
                f'-COUNTBLANK({sn}!D{ind_first}:D{ind_last})',
                f'=COUNTA({sn}!B{ind_first}:B{ind_last})'
                f'-COUNTBLANK({sn}!B{ind_first}:B{ind_last})'
            ))

        secs.append((
            'Casos de Éxito',
            "=COUNTA('CASOS DE ÉXITO'!B4:B13)"
            "-COUNTBLANK('CASOS DE ÉXITO'!B4:B13)",
            '10'
        ))
        secs.append((
            'Metas Generales',
            "=COUNTA('METAS GENERALES'!D4:D18)"
            "-COUNTBLANK('METAS GENERALES'!D4:D18)",
            "=COUNTA('METAS GENERALES'!C4:C18)"
            "-COUNTBLANK('METAS GENERALES'!C4:C18)"
        ))

        for i, (nm, f_a, f_t) in enumerate(secs):
            r = 5 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=nm),
                          font=self.E.FNT_LBL, fill=self.E.FL_W)
            ws.cell(row=r, column=2, value=f_a)
            self.E.aplicar(ws.cell(row=r, column=2),
                          font=self.E.FNT_FRM, fill=self.E.FL_LCK,
                          align=self.E.ALN_C)
            ws.cell(row=r, column=3, value=f_t)
            self.E.aplicar(ws.cell(row=r, column=3),
                          font=self.E.FNT_FRM, fill=self.E.FL_LCK,
                          align=self.E.ALN_C)
            ws.cell(row=r, column=4,
                    value=f'=IF(C{r}>0,B{r}/C{r},0)')
            self.E.aplicar(
                ws.cell(row=r, column=4),
                font=Font(name='Calibri', size=11, bold=True,
                         color='1B3A5C'),
                fill=self.E.FL_LCK, align=self.E.ALN_C
            )
            ws.cell(row=r, column=4).number_format = '0%'

        lr = 5 + len(secs) - 1
        ws.conditional_formatting.add(
            f'D5:D{lr}',
            CellIsRule(operator='greaterThanOrEqual', formula=['0.8'],
                      fill=self.E.FL_GRN))
        ws.conditional_formatting.add(
            f'D5:D{lr}',
            CellIsRule(operator='between', formula=['0.4', '0.79'],
                      fill=PatternFill('solid', fgColor='FEF9E7')))
        ws.conditional_formatting.add(
            f'D5:D{lr}',
            CellIsRule(operator='lessThan', formula=['0.4'],
                      fill=self.E.FL_RED))

        tr = lr + 2
        ws.merge_cells(f'A{tr}:C{tr}')
        ws.cell(row=tr, column=1, value='COMPLETITUD GLOBAL').font = \
            Font(name='Calibri', size=12, bold=True, color='1B3A5C')
        ws.cell(row=tr, column=4, value=f'=AVERAGE(D5:D{lr})')
        ws.cell(row=tr, column=4).font = \
            Font(name='Calibri', size=14, bold=True, color='1B3A5C')
        ws.cell(row=tr, column=4).number_format = '0%'
        ws.cell(row=tr, column=4).alignment = self.E.ALN_C

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password



# ═══════════════════════════════════════════════════════════════════════
# CONSTRUCTOR EXCEL — TERRITORIAL
# ═══════════════════════════════════════════════════════════════════════

class ConstructorExcelTerritorial:
    """
    Genera la matriz Excel de captura para unidades TERRITORIALES
    (Hub, Oficom, OfiReg).

    Estructura: 22 hojas
    - INSTRUCCIONES, CONFIGURACIÓN, TENDENCIAS POR EJE, DOFA
    - 15 hojas de contribución a ejes (MP×4 + TUR×4 + INV×3 + EXP×4)
    - METAS GENERALES, PRESUPUESTO, MONITOREO
    """

    def __init__(self, config: Config):
        self.cfg = config
        self.E = EstilosExcel

    def generar(
        self,
        nombre_unidad: str,
        tipo_unidad: str,
        trimestre: str = 'Q1',
        anio: str = '2026',
        num_lineas: int = 0,  # No aplica para territorial
        datos_base: Optional[Dict] = None,
        output_dir: Optional[str] = None
    ) -> str:
        """Genera Excel territorial completo."""
        output_dir = output_dir or self.cfg.dir_entrada
        wb = openpyxl.Workbook()
        datos = datos_base or {}

        self._crear_instrucciones(wb, nombre_unidad, tipo_unidad)
        self._crear_configuracion(wb, nombre_unidad, tipo_unidad,
                                  trimestre, anio)
        self._crear_tendencias_por_eje(wb, nombre_unidad, datos)
        self._crear_dofa(wb, nombre_unidad, datos)

        # Hojas de contribución a cada eje misional
        for eje_key in ORDEN_EJES:
            eje = EJES_REFERENCIA[eje_key]
            for le_idx, le_nombre in enumerate(eje.lineas, 1):
                contrib_data = None
                if 'contribuciones' in datos:
                    contrib_list = datos['contribuciones'].get(eje_key, [])
                    if le_idx <= len(contrib_list):
                        contrib_data = contrib_list[le_idx - 1]
                self._crear_hoja_contribucion(
                    wb, nombre_unidad, eje, le_idx, le_nombre,
                    contrib_data
                )

        self._crear_metas(wb, nombre_unidad, datos)
        self._crear_presupuesto(wb, nombre_unidad)
        self._crear_monitoreo(wb, nombre_unidad)

        # Guardar
        safe = _safe_filename(nombre_unidad)
        filename = f'{trimestre}_{anio}_{tipo_unidad}_{safe}.xlsx'
        path = os.path.join(output_dir, filename)
        wb.save(path)

        n_hojas = len(wb.sheetnames)
        log.info(f"✅ Excel territorial: {filename} ({n_hojas} hojas)")
        return path

    def _crear_instrucciones(self, wb, nombre: str, tipo: str):
        ws = wb.active
        ws.title = 'INSTRUCCIONES'
        ws.sheet_properties.tabColor = '1B3A5C'
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 70

        ws.merge_cells('B2:C2')
        ws['B2'].value = f'MATRIZ DE CAPTURA — {tipo}'
        ws['B2'].font = Font(name='Calibri', size=18, bold=True,
                            color='1B3A5C')
        ws.merge_cells('B3:C3')
        ws['B3'].value = f'{nombre} — ProColombia'
        ws['B3'].font = Font(name='Calibri', size=12, color='666666')

        instrucciones = [
            ('¿Qué es esta matriz?',
             'Formato oficial para reportar avances trimestrales. '
             f'Como {tipo}, usted reporta su contribución a cada '
             'eje misional (Marca País, Turismo, Inversión, Exportaciones).'),
            ('Hojas a diligenciar',
             '1) CONFIGURACIÓN → Verifique datos\n'
             '2) TENDENCIAS POR EJE → Tendencias/Foco/Aporte por eje\n'
             '3) DOFA → Estado de cada ítem\n'
             '4-18) CONTRIBUCIONES → Sus acciones por cada línea\n'
             '19) METAS → Avance acumulado\n'
             '20) PRESUPUESTO → Cifras estimadas'),
            ('Código de colores',
             '🔵 Azul = Usted diligencia\n'
             '🟡 Amarillo = Info base (no editar)\n'
             '⬜ Gris = Automático\n'
             '🔴 Rojo = ¡Pendiente!'),
        ]
        for i, (t, d) in enumerate(instrucciones):
            r = 5 + i * 3
            ws.cell(row=r, column=2, value=t).font = self.E.FNT_H2
            ws.merge_cells(start_row=r, start_column=3,
                          end_row=r + 1, end_column=3)
            c = ws.cell(row=r, column=3, value=d)
            c.font = self.E.FNT_N
            c.alignment = self.E.ALN_W
        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_configuracion(self, wb, nombre, tipo, trimestre, anio):
        ws = wb.create_sheet('CONFIGURACIÓN')
        ws.sheet_properties.tabColor = '2E6DA4'
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40

        ws.merge_cells('A1:B1')
        ws['A1'].value = 'CONFIGURACIÓN GENERAL'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR

        campos = [
            ('Trimestre en seguimiento', trimestre),
            ('Año', anio),
            ('Tipo de unidad', tipo),
            ('Nombre de la unidad', nombre),
            ('Fecha de envío lineamientos', ''),
            ('Fecha límite diligenciamiento', ''),
        ]
        for i, (lbl, val) in enumerate(campos):
            r = 3 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=lbl),
                          font=self.E.FNT_LBL, fill=self.E.FL_LCK)
            self.E.aplicar(ws.cell(row=r, column=2, value=val),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)

        dv_q = DataValidation(type='list', formula1='"Q1,Q2,Q3,Q4"')
        ws.add_data_validation(dv_q)
        dv_q.add('B3')
        dv_t = DataValidation(type='list',
                              formula1='"HUB,OFICOM,OFIREG"')
        ws.add_data_validation(dv_t)
        dv_t.add('B5')
        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_tendencias_por_eje(self, wb, nombre: str, datos: Dict):
        """Crea la hoja de tendencias con 3 secciones (TUR, INV, EXP)."""
        ws = wb.create_sheet('TENDENCIAS POR EJE')
        ws.sheet_properties.tabColor = '8E44AD'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 65
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 65

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'TENDENCIAS DESDE {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR
        ws.merge_cells('A2:D2')
        ws['A2'].value = ('Para cada eje misional: registre Tendencias, '
                          'Foco/Prioridades y Aporte de su unidad.')
        ws['A2'].font = self.E.FNT_SM

        tend_data = datos.get('tendencias_por_eje', {})
        r = 4
        for eje_key in ['TUR', 'INV', 'EXP']:
            eje = EJES_REFERENCIA[eje_key]

            # Encabezado del eje
            ws.merge_cells(start_row=r, start_column=1,
                          end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value=eje.nombre)
            c.font = self.E.FNT_WB
            c.fill = PatternFill('solid', fgColor=eje.color)
            c.alignment = self.E.ALN_C
            r += 1

            eje_tend = tend_data.get(eje_key, {})
            secciones = [
                ('Tendencias', 'tendencias', 5),
                ('Foco / Prioridades', 'foco', 5),
                ('Aporte de la unidad', 'aporte', 5),
            ]
            for sec_nombre, sec_key, max_items in secciones:
                ws.merge_cells(start_row=r, start_column=1,
                              end_row=r, end_column=1)
                self.E.aplicar(ws.cell(row=r, column=1, value=sec_nombre),
                              font=self.E.FNT_LBL,
                              fill=PatternFill('solid', fgColor='D5D8DC'))
                ws.merge_cells(start_row=r, start_column=2,
                              end_row=r, end_column=4)
                items = eje_tend.get(sec_key, [])
                texto = '\n'.join(f'• {it}' for it in items) if items else ''
                self.E.aplicar(ws.cell(row=r, column=2, value=texto),
                              fill=self.E.FL_EDT, prot=self.E.PROT_U)
                ws.row_dimensions[r].height = 80
                r += 1
            r += 1  # Separador

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_dofa(self, wb, nombre: str, datos: Dict):
        """Crea DOFA territorial — maneja formato dict con estado."""
        ws = wb.create_sheet('DOFA')
        ws.sheet_properties.tabColor = 'C0392B'
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 55

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'SEGUIMIENTO DOFA — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR
        ws.merge_cells('A2:D2')
        ws['A2'].value = ('Col B = Hoja de Ruta original. '
                          'Col C = Estado. Col D = Actualización.')
        ws['A2'].font = self.E.FNT_SM

        for c, h in enumerate(
            ['ID', 'Información Hoja de Ruta (Base)',
             'Estado Seguimiento', 'Actualización / Cambio'], 1
        ):
            self.E.aplicar(ws.cell(row=4, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        dv_d = DataValidation(
            type='list',
            formula1='"Se mantiene,Se elimina,Se actualiza"',
            allow_blank=True
        )
        ws.add_data_validation(dv_d)

        cuads = ['DEBILIDADES', 'OPORTUNIDADES', 'FORTALEZAS', 'AMENAZAS']
        fills_c = {'DEB': 'C0392B', 'OPO': '2E6DA4',
                   'FOR': '27AE60', 'AME': 'F39C12'}
        n = self.cfg.max_dofa_por_cuadrante

        r = 5
        for cq in cuads:
            pre = cq[:3]
            ws.merge_cells(start_row=r, start_column=1,
                          end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value=cq)
            c.font = self.E.FNT_WB
            c.fill = PatternFill('solid', fgColor=fills_c[pre])
            c.alignment = self.E.ALN_C
            r += 1

            items_base = datos.get('dofa', {}).get(cq, [])

            for i in range(n):
                self.E.aplicar(ws.cell(row=r, column=1, value=f'{pre}-{i+1}'),
                              font=self.E.FNT_SM, fill=self.E.FL_LCK)
                # Handle both string and dict formats
                item = items_base[i] if i < len(items_base) else None
                if isinstance(item, dict):
                    base_text = item.get('base', '')
                    estado = item.get('estado', '')
                    actualizacion = item.get('actualizacion', '')
                elif isinstance(item, str):
                    base_text, estado, actualizacion = item, '', ''
                else:
                    base_text, estado, actualizacion = '', '', ''

                self.E.aplicar(ws.cell(row=r, column=2, value=base_text),
                              fill=self.E.FL_BAS)
                self.E.aplicar(ws.cell(row=r, column=3, value=estado),
                              fill=self.E.FL_EDT, prot=self.E.PROT_U)
                dv_d.add(f'C{r}')
                self.E.aplicar(ws.cell(row=r, column=4, value=actualizacion),
                              fill=self.E.FL_EDT, prot=self.E.PROT_U)
                r += 1

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_hoja_contribucion(
        self, wb, nombre_unidad: str,
        eje: EjeMisional, le_idx: int, le_nombre: str,
        contrib_data: Optional[Dict]
    ):
        """Crea una hoja de contribución a una línea de un eje."""
        sn = f'{eje.prefijo}_LE{le_idx}'
        ws = wb.create_sheet(sn)
        ws.sheet_properties.tabColor = eje.color

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 20

        # Encabezado con nombre del eje y línea
        ws.merge_cells('A1:E1')
        ws['A1'].value = (f'CONTRIBUCIÓN A {eje.nombre} — '
                          f'LÍNEA {le_idx}')
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = PatternFill('solid', fgColor=eje.color)

        # Nombre de la línea (bloqueado)
        ws.merge_cells('A2:E2')
        ws['A2'].value = f'{le_idx}. {_trunc(le_nombre, 300)}'
        ws['A2'].font = Font(name='Calibri', size=10, bold=True,
                            color='1B3A5C')
        ws['A2'].alignment = self.E.ALN_W

        # Origen
        ws.merge_cells('A3:E3')
        ws['A3'].value = f'Contribución de: {nombre_unidad}'
        ws['A3'].font = self.E.FNT_SM

        # Sección acciones
        ws.merge_cells('A5:E5')
        ws['A5'].value = f'ACCIONES DE {nombre_unidad.upper()}'
        ws['A5'].font = self.E.FNT_WB
        ws['A5'].fill = self.E.FL_SUB

        headers = ['#', 'Acción',
                   'Actividades Clave',
                   'Avance del Q', 'Estado']
        for c, h in enumerate(headers, 1):
            self.E.aplicar(
                ws.cell(row=6, column=c, value=h),
                font=self.E.FNT_WB,
                fill=PatternFill('solid', fgColor='34495E'),
                align=self.E.ALN_C
            )

        dv_ea = DataValidation(
            type='list',
            formula1='"Completada,En progreso,Sin avance,No aplica"',
            allow_blank=True
        )
        ws.add_data_validation(dv_ea)

        max_acc = self.cfg.max_acciones_por_linea
        acciones = contrib_data.get('acciones', []) if contrib_data else []

        for i in range(max_acc):
            r = 7 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_SM, fill=self.E.FL_LCK)
            acc_text = acciones[i].get('accion', '') \
                if i < len(acciones) else ''
            act_text = acciones[i].get('actividad', '') \
                if i < len(acciones) else ''
            self.E.aplicar(ws.cell(row=r, column=2, value=acc_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=3, value=act_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=5),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U,
                          align=self.E.ALN_C)
            dv_ea.add(f'E{r}')

        # Formato condicional
        for rx in range(7, 7 + max_acc):
            ws.conditional_formatting.add(
                f'D{rx}',
                FormulaRule(
                    formula=[f'AND(NOT(ISBLANK(B{rx})),ISBLANK(D{rx}))'],
                    fill=self.E.FL_RED
                ))

        # Indicadores
        ir = 7 + max_acc + 1
        ws.merge_cells(f'A{ir}:E{ir}')
        ws.cell(row=ir, column=1, value='INDICADORES DE ÉXITO')
        ws.cell(row=ir, column=1).font = self.E.FNT_WB
        ws.cell(row=ir, column=1).fill = self.E.FL_SUB

        ih = ir + 1
        for c, h in enumerate(
            ['#', 'Indicador', 'Meta', 'Avance', 'Observaciones'], 1
        ):
            self.E.aplicar(
                ws.cell(row=ih, column=c, value=h),
                font=self.E.FNT_WB,
                fill=PatternFill('solid', fgColor='34495E'),
                align=self.E.ALN_C
            )

        inds = contrib_data.get('indicadores', []) if contrib_data else []
        max_ind = self.cfg.max_indicadores_por_linea
        for i in range(max_ind):
            r = ih + 1 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_SM, fill=self.E.FL_LCK)
            ind_text = inds[i].get('indicador', '') \
                if i < len(inds) else ''
            meta_text = inds[i].get('meta', '') \
                if i < len(inds) else ''
            self.E.aplicar(ws.cell(row=r, column=2, value=ind_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=3, value=meta_text),
                          fill=self.E.FL_BAS)
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=5),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_metas(self, wb, nombre: str, datos: Dict):
        """Reutiliza metas del misional."""
        ConstructorExcelMisional(self.cfg)._crear_metas(wb, nombre, datos)

    def _crear_presupuesto(self, wb, nombre: str):
        ws = wb.create_sheet('PRESUPUESTO')
        ws.sheet_properties.tabColor = 'D4AC0D'
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'PRESUPUESTO ESTIMADO — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR

        for c, h in enumerate(['#', 'Concepto', 'Monto (COP)', 'Notas'], 1):
            self.E.aplicar(ws.cell(row=3, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)
        for i in range(10):
            r = 4 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=i + 1),
                          font=self.E.FNT_SM, fill=self.E.FL_LCK)
            self.E.aplicar(ws.cell(row=r, column=2),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=3),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
            self.E.aplicar(ws.cell(row=r, column=4),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_monitoreo(self, wb, nombre: str):
        ws = wb.create_sheet('MONITOREO')
        ws.sheet_properties.tabColor = '2C3E50'
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'MONITOREO — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR
        ws['A2'].value = 'Cálculo automático. No editar.'
        ws['A2'].font = self.E.FNT_SM

        for c, h in enumerate(
            ['Sección', 'Completados', 'Esperados', '% Completitud'], 1
        ):
            self.E.aplicar(ws.cell(row=4, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        acc_first, acc_last = 7, 6 + self.cfg.max_acciones_por_linea

        secs = []
        for eje_key in ORDEN_EJES:
            eje = EJES_REFERENCIA[eje_key]
            for le_idx in range(1, len(eje.lineas) + 1):
                sn = f"'{eje.prefijo}_LE{le_idx}'"
                secs.append((
                    f'{eje.nombre} — LE{le_idx} Avances',
                    f'=COUNTA({sn}!D{acc_first}:D{acc_last})'
                    f'-COUNTBLANK({sn}!D{acc_first}:D{acc_last})',
                    f'=COUNTA({sn}!B{acc_first}:B{acc_last})'
                    f'-COUNTBLANK({sn}!B{acc_first}:B{acc_last})'
                ))

        secs.append((
            'Metas Generales',
            "=COUNTA('METAS GENERALES'!D4:D18)"
            "-COUNTBLANK('METAS GENERALES'!D4:D18)",
            "=COUNTA('METAS GENERALES'!C4:C18)"
            "-COUNTBLANK('METAS GENERALES'!C4:C18)"
        ))

        for i, (nm, f_a, f_t) in enumerate(secs):
            r = 5 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=nm),
                          font=self.E.FNT_LBL, fill=self.E.FL_W)
            ws.cell(row=r, column=2, value=f_a)
            self.E.aplicar(ws.cell(row=r, column=2),
                          font=self.E.FNT_FRM, fill=self.E.FL_LCK,
                          align=self.E.ALN_C)
            ws.cell(row=r, column=3, value=f_t)
            self.E.aplicar(ws.cell(row=r, column=3),
                          font=self.E.FNT_FRM, fill=self.E.FL_LCK,
                          align=self.E.ALN_C)
            ws.cell(row=r, column=4,
                    value=f'=IF(C{r}>0,B{r}/C{r},0)')
            self.E.aplicar(ws.cell(row=r, column=4),
                          font=Font(name='Calibri', size=11, bold=True,
                                   color='1B3A5C'),
                          fill=self.E.FL_LCK, align=self.E.ALN_C)
            ws.cell(row=r, column=4).number_format = '0%'

        lr = 5 + len(secs) - 1
        ws.conditional_formatting.add(
            f'D5:D{lr}',
            CellIsRule(operator='greaterThanOrEqual', formula=['0.8'],
                      fill=self.E.FL_GRN))
        ws.conditional_formatting.add(
            f'D5:D{lr}',
            CellIsRule(operator='lessThan', formula=['0.4'],
                      fill=self.E.FL_RED))

        tr = lr + 2
        ws.merge_cells(f'A{tr}:C{tr}')
        ws.cell(row=tr, column=1, value='COMPLETITUD GLOBAL').font = \
            Font(name='Calibri', size=12, bold=True, color='1B3A5C')
        ws.cell(row=tr, column=4, value=f'=AVERAGE(D5:D{lr})')
        ws.cell(row=tr, column=4).font = \
            Font(name='Calibri', size=14, bold=True, color='1B3A5C')
        ws.cell(row=tr, column=4).number_format = '0%'
        ws.cell(row=tr, column=4).alignment = self.E.ALN_C

        ws.protection.sheet = True
        ws.protection.password = self.cfg.password



# ═══════════════════════════════════════════════════════════════════════
# DATOS DE EJEMPLO — HUB NORTEAMÉRICA
# ═══════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════
# CONSTRUCTOR EXCEL — TRANSVERSAL
# ═══════════════════════════════════════════════════════════════════════

class ConstructorExcelTransversal:
    """
    Genera la matriz Excel para unidades TRANSVERSALES.
    Combina: líneas estratégicas propias (como misional)
    + contribuciones a 4 ejes (como territorial).
    """

    def __init__(self, config: Config):
        self.cfg = config
        self.E = EstilosExcel
        self._misional = ConstructorExcelMisional(config)
        self._territorial = ConstructorExcelTerritorial(config)

    def generar(
        self,
        nombre_unidad: str,
        tipo_unidad: str = 'TRANSVERSAL',
        trimestre: str = 'Q1',
        anio: str = '2026',
        num_lineas: int = 3,
        datos_base: Optional[Dict] = None,
        output_dir: Optional[str] = None,
    ) -> str:
        """Genera Excel transversal completo."""
        num_lineas = max(1, min(num_lineas, self.cfg.max_lineas_estrategicas))
        output_dir = output_dir or self.cfg.dir_entrada
        datos = datos_base or {}
        wb = openpyxl.Workbook()

        # 1-2: Instrucciones + Configuración
        self._crear_instrucciones(wb, nombre_unidad, num_lineas)
        self._crear_configuracion(wb, nombre_unidad, tipo_unidad,
                                  trimestre, anio)

        # 3: DOFA (reutiliza territorial que maneja dicts)
        self._territorial._crear_dofa(wb, nombre_unidad, datos)

        # 4-8: Líneas estratégicas propias
        for le in range(1, num_lineas + 1):
            self._misional._crear_linea_estrategica(
                wb, nombre_unidad, le, datos
            )

        # 9-23: Contribuciones a ejes misionales
        for eje_key in ORDEN_EJES:
            eje = EJES_REFERENCIA[eje_key]
            for le_idx, le_nombre in enumerate(eje.lineas, 1):
                contrib_data = None
                if 'contribuciones' in datos:
                    contrib_list = datos['contribuciones'].get(eje_key, [])
                    if le_idx <= len(contrib_list):
                        contrib_data = contrib_list[le_idx - 1]
                self._territorial._crear_hoja_contribucion(
                    wb, nombre_unidad, eje, le_idx, le_nombre, contrib_data
                )

        # 24-26: Metas, Presupuesto, Monitoreo
        self._misional._crear_metas(wb, nombre_unidad, datos)
        self._territorial._crear_presupuesto(wb, nombre_unidad)
        self._crear_monitoreo(wb, nombre_unidad, num_lineas)

        # Guardar
        safe = _safe_filename(nombre_unidad)
        filename = f'{trimestre}_{anio}_{tipo_unidad}_{safe}.xlsx'
        path = os.path.join(output_dir, filename)
        wb.save(path)
        log.info(f"✅ Excel transversal: {filename} ({len(wb.sheetnames)} hojas)")
        return path

    def _crear_instrucciones(self, wb, nombre: str, num_lineas: int):
        ws = wb.active
        ws.title = 'INSTRUCCIONES'
        ws.sheet_properties.tabColor = '1B3A5C'
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 70

        ws.merge_cells('B2:C2')
        ws['B2'].value = 'MATRIZ DE CAPTURA — TRANSVERSAL'
        ws['B2'].font = Font(name='Calibri', size=18, bold=True,
                            color='1B3A5C')
        ws.merge_cells('B3:C3')
        ws['B3'].value = f'{nombre} — ProColombia'
        ws['B3'].font = Font(name='Calibri', size=12, color='666666')

        instrucciones = [
            ('¿Qué es esta matriz?',
             f'Como área transversal, usted reporta:\n'
             f'(A) Sus {num_lineas} líneas estratégicas propias\n'
             f'(B) Su contribución a cada línea de los 4 ejes misionales\n'
             '(Marca País, Turismo, Inversión, Exportaciones)'),
            ('Hojas a diligenciar',
             '1) CONFIGURACIÓN → Verifique datos\n'
             '2) DOFA → Estado de cada ítem\n'
             f'3) LÍNEAS PROPIAS (1 a {num_lineas}) → Sus avances\n'
             '4) CONTRIBUCIONES (MP, TUR, INV, EXP) → Sus acciones por línea\n'
             '5) METAS → Avance acumulado\n'
             '6) PRESUPUESTO → Cifras estimadas'),
            ('Código de colores',
             '🔵 Azul = Usted diligencia\n'
             '🟡 Amarillo = Info base (no editar)\n'
             '⬜ Gris = Automático'),
        ]
        for i, (t, d) in enumerate(instrucciones):
            r = 5 + i * 3
            ws.cell(row=r, column=2, value=t).font = self.E.FNT_H2
            ws.merge_cells(start_row=r, start_column=3,
                          end_row=r + 1, end_column=3)
            c = ws.cell(row=r, column=3, value=d)
            c.font = self.E.FNT_N
            c.alignment = self.E.ALN_W
        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_configuracion(self, wb, nombre, tipo, trimestre, anio):
        ws = wb.create_sheet('CONFIGURACIÓN')
        ws.sheet_properties.tabColor = '2E6DA4'
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40
        ws.merge_cells('A1:B1')
        ws['A1'].value = 'CONFIGURACIÓN GENERAL'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR
        for i, (lbl, val) in enumerate([
            ('Trimestre en seguimiento', trimestre), ('Año', anio),
            ('Tipo de unidad', tipo), ('Nombre de la unidad', nombre),
            ('Fecha de envío lineamientos', ''),
            ('Fecha límite diligenciamiento', ''),
        ]):
            r = 3 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=lbl),
                          font=self.E.FNT_LBL, fill=self.E.FL_LCK)
            self.E.aplicar(ws.cell(row=r, column=2, value=val),
                          fill=self.E.FL_EDT, prot=self.E.PROT_U)
        ws.protection.sheet = True
        ws.protection.password = self.cfg.password

    def _crear_monitoreo(self, wb, nombre: str, num_lineas: int):
        ws = wb.create_sheet('MONITOREO')
        ws.sheet_properties.tabColor = '2C3E50'
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        ws.merge_cells('A1:D1')
        ws['A1'].value = f'MONITOREO — {nombre}'
        ws['A1'].font = self.E.FNT_H
        ws['A1'].fill = self.E.FL_HDR
        ws['A2'].value = 'Cálculo automático. No editar.'
        ws['A2'].font = self.E.FNT_SM

        for c, h in enumerate(
            ['Sección', 'Completados', 'Esperados', '% Completitud'], 1
        ):
            self.E.aplicar(ws.cell(row=4, column=c, value=h),
                          font=self.E.FNT_WB, fill=self.E.FL_SUB,
                          align=self.E.ALN_C)

        acc_f, acc_l = 6, 5 + self.cfg.max_acciones_por_linea
        ind_f = acc_l + 3
        ind_l = ind_f + self.cfg.max_indicadores_por_linea - 1
        contrib_f, contrib_l = 7, 6 + self.cfg.max_acciones_por_linea

        secs = []
        # Líneas propias
        for n in range(1, num_lineas + 1):
            sn = f"'LÍNEA ESTRATÉGICA {n}'"
            secs.append((f'Línea propia {n} — Avances',
                         f'=COUNTA({sn}!D{acc_f}:D{acc_l})'
                         f'-COUNTBLANK({sn}!D{acc_f}:D{acc_l})',
                         f'=COUNTA({sn}!B{acc_f}:B{acc_l})'
                         f'-COUNTBLANK({sn}!B{acc_f}:B{acc_l})'))

        # Contribuciones
        for eje_key in ORDEN_EJES:
            eje = EJES_REFERENCIA[eje_key]
            for le_idx in range(1, len(eje.lineas) + 1):
                sn = f"'{eje.prefijo}_LE{le_idx}'"
                secs.append((
                    f'{eje.nombre} — LE{le_idx}',
                    f'=COUNTA({sn}!D{contrib_f}:D{contrib_l})'
                    f'-COUNTBLANK({sn}!D{contrib_f}:D{contrib_l})',
                    f'=COUNTA({sn}!B{contrib_f}:B{contrib_l})'
                    f'-COUNTBLANK({sn}!B{contrib_f}:B{contrib_l})'))

        secs.append(('Metas',
                     "=COUNTA('METAS GENERALES'!D4:D18)"
                     "-COUNTBLANK('METAS GENERALES'!D4:D18)",
                     "=COUNTA('METAS GENERALES'!C4:C18)"
                     "-COUNTBLANK('METAS GENERALES'!C4:C18)"))

        for i, (nm, f_a, f_t) in enumerate(secs):
            r = 5 + i
            self.E.aplicar(ws.cell(row=r, column=1, value=nm),
                          font=self.E.FNT_LBL, fill=self.E.FL_W)
            ws.cell(row=r, column=2, value=f_a)
            self.E.aplicar(ws.cell(row=r, column=2),
                          font=self.E.FNT_FRM, fill=self.E.FL_LCK,
                          align=self.E.ALN_C)
            ws.cell(row=r, column=3, value=f_t)
            self.E.aplicar(ws.cell(row=r, column=3),
                          font=self.E.FNT_FRM, fill=self.E.FL_LCK,
                          align=self.E.ALN_C)
            ws.cell(row=r, column=4, value=f'=IF(C{r}>0,B{r}/C{r},0)')
            self.E.aplicar(ws.cell(row=r, column=4),
                          font=Font(name='Calibri', size=11, bold=True,
                                   color='1B3A5C'),
                          fill=self.E.FL_LCK, align=self.E.ALN_C)
            ws.cell(row=r, column=4).number_format = '0%'

        lr = 5 + len(secs) - 1
        ws.conditional_formatting.add(f'D5:D{lr}',
            CellIsRule(operator='greaterThanOrEqual', formula=['0.8'],
                      fill=self.E.FL_GRN))
        ws.conditional_formatting.add(f'D5:D{lr}',
            CellIsRule(operator='lessThan', formula=['0.4'],
                      fill=self.E.FL_RED))

        tr = lr + 2
        ws.merge_cells(f'A{tr}:C{tr}')
        ws.cell(row=tr, column=1, value='COMPLETITUD GLOBAL').font = \
            Font(name='Calibri', size=12, bold=True, color='1B3A5C')
        ws.cell(row=tr, column=4, value=f'=AVERAGE(D5:D{lr})')
        ws.cell(row=tr, column=4).font = \
            Font(name='Calibri', size=14, bold=True, color='1B3A5C')
        ws.cell(row=tr, column=4).number_format = '0%'
        ws.cell(row=tr, column=4).alignment = self.E.ALN_C
        ws.protection.sheet = True
        ws.protection.password = self.cfg.password


