# -*- coding: utf-8 -*-
"""
Utilidades compartidas: funciones helper, estilos Excel, logging.

Este módulo solo depende de config.py (indirectamente, a través
de quienes lo importan).  No depende de ningún otro módulo del paquete.
"""

from __future__ import annotations

import logging
import re
import time
from datetime import timedelta
from textwrap import shorten
from typing import Any

import pandas as pd
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection,
)

# ═══════════════════════════════════════════════════════════════════════
# LOGGING
# ═══════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger('ProColombia')


# ═══════════════════════════════════════════════════════════════════════
# FUNCIONES HELPER
# ═══════════════════════════════════════════════════════════════════════

def _ss(val: Any, default: str = '') -> str:
    """Safe string: convierte valor a string limpio."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    return str(val).strip()


def _trunc(texto: str, mx: int = 450) -> str:
    """Trunca texto a un máximo de caracteres."""
    if not texto or len(texto) <= mx:
        return texto or ''
    return shorten(texto, width=mx, placeholder='...')


def _safe_filename(nombre: str) -> str:
    """Convierte nombre a formato seguro para archivos."""
    return re.sub(r'[^\w\s-]', '', nombre).strip().replace(' ', '_')


def medir_tiempo(nombre: str = "") -> callable:
    """Crea un medidor de tiempo reutilizable."""
    inicio = time.time()
    def fin():
        t = str(timedelta(seconds=time.time() - inicio))
        s = f" para {nombre}" if nombre else ""
        print(f"⏱️  Tiempo transcurrido{s}: {t}")
    return fin


# ═══════════════════════════════════════════════════════════════════════
# ESTILOS EXCEL (reutilizables entre las 3 familias)
# ═══════════════════════════════════════════════════════════════════════

class EstilosExcel:
    """Definición centralizada de estilos para Excel. Reutilizable."""

    # Fuentes
    FNT_H   = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    FNT_H2  = Font(name='Calibri', size=11, bold=True, color='1B3A5C')
    FNT_LBL = Font(name='Calibri', size=10, bold=True, color='333333')
    FNT_N   = Font(name='Calibri', size=10, color='333333')
    FNT_SM  = Font(name='Calibri', size=9, color='666666', italic=True)
    FNT_FRM = Font(name='Calibri', size=10, color='27AE60', bold=True)
    FNT_WB  = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

    # Rellenos
    FL_HDR = PatternFill('solid', fgColor='1B3A5C')
    FL_SUB = PatternFill('solid', fgColor='2E6DA4')
    FL_EDT = PatternFill('solid', fgColor='E8F4FD')
    FL_LCK = PatternFill('solid', fgColor='F2F2F2')
    FL_BAS = PatternFill('solid', fgColor='FFF9E6')
    FL_GRN = PatternFill('solid', fgColor='D5F5E3')
    FL_RED = PatternFill('solid', fgColor='FADBD8')
    FL_W   = PatternFill('solid', fgColor='FFFFFF')

    # Alineación
    ALN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALN_W = Alignment(wrap_text=True, vertical='top')

    # Borde
    BRD = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7'),
    )

    # Protección
    PROT_L = Protection(locked=True)
    PROT_U = Protection(locked=False)

    # Colores de pestañas por línea estratégica
    TAB_COLORS = ['E74C3C', '3498DB', '27AE60', 'F39C12', '8E44AD']

    @classmethod
    def aplicar(cls, cell, font=None, fill=None, align=None, prot=None):
        """Aplica estilos a una celda con defaults."""
        cell.font = font or cls.FNT_N
        cell.fill = fill or cls.FL_W
        cell.alignment = align or cls.ALN_W
        cell.border = cls.BRD
        cell.protection = prot or cls.PROT_L
