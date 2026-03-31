# -*- coding: utf-8 -*-
"""
Tests automatizados para el paquete procolombia.

Ejecutar desde notebook:
    %run procolombia/tests.py

Ejecutar desde terminal:
    python -m procolombia.tests

Cubre:
    • Config: validación de parámetros
    • FamiliaUnidad: mapeo tipo → familia
    • Constructores: generación de Excel de las 3 familias
    • Lectores: lectura de Excel generados (roundtrip)
    • OrquestadorUniversal: detección de familia, pipeline completo
    • Consolidador: generación de Excel maestro
    • Utilidades: _ss, _trunc, _safe_filename, _safe_col, _es_numero
    • Regla de oro: cero iterrows en lectores
"""

from __future__ import annotations

import inspect
import os
import sys
import tempfile
import shutil
from pathlib import Path

# ── Setup de imports ─────────────────────────────────────────────
# Buscar el paquete en el directorio padre de este archivo
_pkg_dir = str(Path(__file__).resolve().parent.parent)
if _pkg_dir not in sys.path:
    sys.path.insert(0, _pkg_dir)


# ═══════════════════════════════════════════════════════════════════════
# FRAMEWORK DE TESTS (mínimo, sin dependencias externas)
# ═══════════════════════════════════════════════════════════════════════

_PASSED = 0
_FAILED = 0
_ERRORS: list = []


def _test(nombre: str, condicion: bool, detalle: str = '') -> None:
    """Registra un test pass/fail."""
    global _PASSED, _FAILED
    if condicion:
        _PASSED += 1
    else:
        _FAILED += 1
        msg = f"  ❌ {nombre}"
        if detalle:
            msg += f" — {detalle}"
        _ERRORS.append(msg)
        print(msg)


def _resumen() -> bool:
    """Imprime resumen y retorna True si todo pasó."""
    total = _PASSED + _FAILED
    print(f"\n{'═' * 60}")
    if _FAILED == 0:
        print(f"  ✅ {_PASSED}/{total} tests pasaron — TODO OK")
    else:
        print(f"  ❌ {_FAILED}/{total} tests FALLARON:")
        for e in _ERRORS:
            print(f"    {e}")
    print(f"{'═' * 60}\n")
    return _FAILED == 0


# ═══════════════════════════════════════════════════════════════════════
# TESTS
# ═══════════════════════════════════════════════════════════════════════

def test_config_defaults():
    """Config con valores por defecto."""
    from procolombia.config import Config
    cfg = Config()
    _test("Config.max_lineas default=5", cfg.max_lineas_estrategicas == 5)
    _test("Config.password default", cfg.password == 'planeacion2026')
    _test("Config.dir_entrada default", cfg.dir_entrada == '01_excels_entrada')


def test_config_validacion():
    """Config rechaza valores inválidos."""
    from procolombia.config import Config
    try:
        Config(max_lineas_estrategicas=0)
        _test("Config rechaza max_lineas=0", False, "No lanzó ValueError")
    except ValueError:
        _test("Config rechaza max_lineas=0", True)

    try:
        Config(max_lineas_estrategicas=7)
        _test("Config rechaza max_lineas=7", False, "No lanzó ValueError")
    except ValueError:
        _test("Config rechaza max_lineas=7", True)

    try:
        Config(max_acciones_por_linea=0)
        _test("Config rechaza max_acciones=0", False, "No lanzó ValueError")
    except ValueError:
        _test("Config rechaza max_acciones=0", True)


def test_config_personalizada():
    """Config acepta valores personalizados válidos."""
    from procolombia.config import Config
    cfg = Config(max_lineas_estrategicas=3, password='test_pw')
    _test("Config personalizada lineas=3", cfg.max_lineas_estrategicas == 3)
    _test("Config personalizada password", cfg.password == 'test_pw')


def test_familia_desde_tipo():
    """FamiliaUnidad.desde_tipo mapea correctamente."""
    from procolombia.config import FamiliaUnidad
    mapeos = {
        'EJE': 'MISIONAL', 'MARCA PAÍS': 'MISIONAL',
        'HUB': 'TERRITORIAL', 'OFICOM': 'TERRITORIAL',
        'OFIREG': 'TERRITORIAL', 'TRANSVERSAL': 'TRANSVERSAL',
    }
    for tipo, familia_esperada in mapeos.items():
        resultado = FamiliaUnidad.desde_tipo(tipo)
        _test(f"desde_tipo('{tipo}')={familia_esperada}",
              resultado.value == familia_esperada)

    # Tipo inválido
    try:
        FamiliaUnidad.desde_tipo('INEXISTENTE')
        _test("desde_tipo('INEXISTENTE') lanza ValueError", False)
    except ValueError:
        _test("desde_tipo('INEXISTENTE') lanza ValueError", True)


def test_utilidades():
    """Funciones helper: _ss, _trunc, _safe_filename."""
    from procolombia.utils import _ss, _trunc, _safe_filename

    # _ss
    _test("_ss(None)=''", _ss(None) == '')
    _test("_ss(float('nan'))=''", _ss(float('nan')) == '')
    _test("_ss(' hola ')='hola'", _ss(' hola ') == 'hola')
    _test("_ss(123)='123'", _ss(123) == '123')

    # _trunc
    _test("_trunc corto no trunca", _trunc('abc', 10) == 'abc')
    _test("_trunc largo trunca", len(_trunc('a' * 500, 100)) <= 100)
    _test("_trunc None retorna ''", _trunc(None) == '')

    # _safe_filename
    _test("_safe_filename limpia", _safe_filename('VP Exportaciones') == 'VP_Exportaciones')
    _test("_safe_filename chars especiales",
          _safe_filename('Ger. Int/Com!') == 'Ger_IntCom')


def test_safe_col_y_es_numero():
    """Helpers de excel_lectores."""
    from procolombia.excel_lectores import _safe_col, _es_numero

    row = ('valor0', 'valor1', 'valor2')
    _test("_safe_col(row, 0)", _safe_col(row, 0) == 'valor0')
    _test("_safe_col(row, 2)", _safe_col(row, 2) == 'valor2')
    _test("_safe_col(row, 5) fuera de rango", _safe_col(row, 5) == '')
    _test("_safe_col default", _safe_col(row, 5, 'X') == 'X')

    _test("_es_numero('1')", _es_numero('1') is True)
    _test("_es_numero('3.0')", _es_numero('3.0') is True)
    _test("_es_numero('abc')", _es_numero('abc') is False)
    _test("_es_numero('')", _es_numero('') is False)


def test_ejes_referencia():
    """EJES_REFERENCIA tiene los 4 ejes correctos."""
    from procolombia.config import EJES_REFERENCIA, ORDEN_EJES
    _test("4 ejes definidos", len(EJES_REFERENCIA) == 4)
    _test("ORDEN_EJES correcto", ORDEN_EJES == ['MP', 'TUR', 'INV', 'EXP'])
    for k in ORDEN_EJES:
        eje = EJES_REFERENCIA[k]
        _test(f"Eje {k} tiene líneas", len(eje.lineas) > 0)
        _test(f"Eje {k} tiene prefijo", eje.prefijo == k)


def test_generar_excel_misional():
    """Genera un Excel misional y verifica estructura."""
    from procolombia.config import Config
    from procolombia.excel_constructores import ConstructorExcelMisional
    import openpyxl

    cfg = Config(max_lineas_estrategicas=3)
    d = tempfile.mkdtemp()
    try:
        c = ConstructorExcelMisional(cfg)
        ruta = c.generar('VP Test', 'EJE', 'Q1', '2026', 3, None, d)
        _test("Excel misional existe", os.path.exists(ruta))

        wb = openpyxl.load_workbook(ruta, read_only=True)
        hojas = set(wb.sheetnames)
        wb.close()

        for h in ['INSTRUCCIONES', 'CONFIGURACIÓN', 'DOFA', 'TENDENCIAS',
                   'LÍNEA ESTRATÉGICA 1', 'CASOS DE ÉXITO', 'METAS GENERALES']:
            _test(f"Hoja '{h}' existe", h in hojas, f"Hojas: {hojas}")

        _test("3 líneas creadas", 'LÍNEA ESTRATÉGICA 3' in hojas)
        _test("LE4 no existe (solo 3)", 'LÍNEA ESTRATÉGICA 4' not in hojas)
    finally:
        shutil.rmtree(d)


def test_generar_excel_territorial():
    """Genera un Excel territorial y verifica estructura."""
    from procolombia.config import Config
    from procolombia.excel_constructores import ConstructorExcelTerritorial
    import openpyxl

    cfg = Config()
    d = tempfile.mkdtemp()
    try:
        c = ConstructorExcelTerritorial(cfg)
        ruta = c.generar('Hub Test', 'HUB', 'Q1', '2026', 0, None, d)
        _test("Excel territorial existe", os.path.exists(ruta))

        wb = openpyxl.load_workbook(ruta, read_only=True)
        hojas = set(wb.sheetnames)
        wb.close()

        for h in ['CONFIGURACIÓN', 'DOFA', 'TENDENCIAS POR EJE',
                   'MP_LE1', 'TUR_LE1', 'INV_LE1', 'EXP_LE1']:
            _test(f"Territorial: hoja '{h}'", h in hojas)

        _test("Territorial: NO tiene TENDENCIAS",
              'TENDENCIAS' not in hojas)
    finally:
        shutil.rmtree(d)


def test_generar_excel_transversal():
    """Genera un Excel transversal y verifica estructura híbrida."""
    from procolombia.config import Config
    from procolombia.excel_constructores import ConstructorExcelTransversal
    import openpyxl

    cfg = Config(max_lineas_estrategicas=3)
    d = tempfile.mkdtemp()
    try:
        c = ConstructorExcelTransversal(cfg)
        ruta = c.generar('GIC Test', 'TRANSVERSAL', 'Q1', '2026', 3, None, d)
        _test("Excel transversal existe", os.path.exists(ruta))

        wb = openpyxl.load_workbook(ruta, read_only=True)
        hojas = set(wb.sheetnames)
        wb.close()

        # Tiene líneas propias Y contribuciones
        _test("Transversal: tiene LE1", 'LÍNEA ESTRATÉGICA 1' in hojas)
        _test("Transversal: tiene MP_LE1", 'MP_LE1' in hojas)
        _test("Transversal: tiene DOFA", 'DOFA' in hojas)
    finally:
        shutil.rmtree(d)


def test_roundtrip_misional():
    """Genera Excel misional → lee → verifica datos."""
    from procolombia.config import Config
    from procolombia.excel_constructores import ConstructorExcelMisional
    from procolombia.excel_lectores import LectorExcel

    cfg = Config(max_lineas_estrategicas=3)
    d = tempfile.mkdtemp()
    try:
        ruta = ConstructorExcelMisional(cfg).generar(
            'VP Roundtrip', 'EJE', 'Q2', '2026', 3, None, d)
        data = LectorExcel(cfg).leer(ruta)

        _test("Roundtrip: unidad correcta",
              data['config']['Nombre de la unidad'] == 'VP Roundtrip')
        _test("Roundtrip: trimestre",
              data['config']['Trimestre en seguimiento'] == 'Q2')
        _test("Roundtrip: familia MISIONAL",
              data['familia'] == 'MISIONAL')
        _test("Roundtrip: 3 líneas", len(data['lineas']) == 3)
        _test("Roundtrip: dofa tiene 4 cuadrantes",
              len(data['dofa']) == 4)
        _test("Roundtrip: tendencias es lista",
              isinstance(data['tendencias'], list))
    finally:
        shutil.rmtree(d)


def test_roundtrip_territorial():
    """Genera Excel territorial → lee → verifica datos."""
    from procolombia.config import Config
    from procolombia.excel_constructores import ConstructorExcelTerritorial
    from procolombia.excel_lectores import LectorExcelTerritorial

    cfg = Config()
    d = tempfile.mkdtemp()
    try:
        ruta = ConstructorExcelTerritorial(cfg).generar(
            'Hub RT', 'HUB', 'Q1', '2026', 0, None, d)
        data = LectorExcelTerritorial(cfg).leer(ruta)

        _test("RT Territorial: familia",
              data['familia'] == 'TERRITORIAL')
        _test("RT Territorial: tiene contribuciones",
              'contribuciones' in data)
        _test("RT Territorial: 4 ejes en contrib",
              len(data['contribuciones']) == 4)
        _test("RT Territorial: lineas vacía",
              data['lineas'] == [])
    finally:
        shutil.rmtree(d)


def test_deteccion_familia():
    """OrquestadorUniversal detecta familia correctamente."""
    from procolombia.config import Config, FamiliaUnidad
    from procolombia.orquestador import OrquestadorUniversal

    cfg = Config(max_lineas_estrategicas=3)
    d = tempfile.mkdtemp()
    try:
        orq = OrquestadorUniversal(config=cfg, base_dir=d)
        r_m = orq.generar_excel('VP X', 'EJE', trimestre='Q1', anio='2026', num_lineas=2)
        r_t = orq.generar_excel('Hub X', 'HUB', trimestre='Q1', anio='2026')
        r_tr = orq.generar_excel('GIC X', 'TRANSVERSAL', trimestre='Q1', anio='2026', num_lineas=2)

        _test("Detecta MISIONAL", orq._detectar_familia(r_m) == FamiliaUnidad.MISIONAL)
        _test("Detecta TERRITORIAL", orq._detectar_familia(r_t) == FamiliaUnidad.TERRITORIAL)
        _test("Detecta TRANSVERSAL", orq._detectar_familia(r_tr) == FamiliaUnidad.TRANSVERSAL)
    finally:
        shutil.rmtree(d)


def test_consolidar():
    """OrquestadorUniversal consolida sin errores."""
    from procolombia.config import Config
    from procolombia.orquestador import OrquestadorUniversal

    cfg = Config(max_lineas_estrategicas=2)
    d = tempfile.mkdtemp()
    try:
        orq = OrquestadorUniversal(config=cfg, base_dir=d)
        orq.generar_excel('VP C', 'EJE', trimestre='Q1', anio='2026', num_lineas=2)
        orq.generar_excel('Hub C', 'HUB', trimestre='Q1', anio='2026')

        ruta = orq.consolidar()
        _test("Consolidado generado", ruta is not None and os.path.exists(ruta))
    finally:
        shutil.rmtree(d)


def test_lectura_real_hub():
    """Lee el Excel real de Hub Norteamérica (si está disponible)."""
    from procolombia.config import Config, FamiliaUnidad
    from procolombia.excel_lectores import LectorExcelTerritorial

    real = '/mnt/user-data/uploads/1774679090635_Q1_2026_HUB_Hub_Norteamérica.xlsx'
    if not os.path.exists(real):
        print("  ⏭️  Archivo real no disponible, skip")
        return

    cfg = Config()
    data = LectorExcelTerritorial(cfg).leer(real)
    _test("Real Hub: familia TERRITORIAL",
          data['familia'] == 'TERRITORIAL')
    n_c = sum(len(le.get('acciones', []))
              for les in data['contribuciones'].values() for le in les)
    _test("Real Hub: tiene contribuciones", n_c > 0,
          f"encontradas: {n_c}")


def test_cero_iterrows():
    """Verifica que excel_lectores.py no usa .iterrows() en código."""
    from procolombia import excel_lectores
    src = inspect.getsource(excel_lectores)

    # Buscar .iterrows() real (no en comentarios)
    lineas_codigo = []
    for linea in src.split('\n'):
        stripped = linea.strip()
        # Ignorar comentarios y docstrings
        if stripped.startswith('#') or stripped.startswith("'") or \
           stripped.startswith('"'):
            continue
        if '.iterrows()' in linea:
            lineas_codigo.append(linea.strip())

    _test("CERO .iterrows() en código de lectores",
          len(lineas_codigo) == 0,
          f"Encontrados: {lineas_codigo}")


def test_cero_iterrows_constructores():
    """Verifica que excel_constructores.py no usa .iterrows()."""
    from procolombia import excel_constructores
    src = inspect.getsource(excel_constructores)
    _test("CERO .iterrows() en constructores",
          '.iterrows()' not in src)


# ═══════════════════════════════════════════════════════════════════════
# RUNNER
# ═══════════════════════════════════════════════════════════════════════

def run_all() -> bool:
    """Ejecuta todos los tests y retorna True si pasaron."""
    print("\n" + "═" * 60)
    print("  TESTS AUTOMATIZADOS — procolombia v5.1")
    print("═" * 60 + "\n")

    tests = [
        test_config_defaults,
        test_config_validacion,
        test_config_personalizada,
        test_familia_desde_tipo,
        test_utilidades,
        test_safe_col_y_es_numero,
        test_ejes_referencia,
        test_generar_excel_misional,
        test_generar_excel_territorial,
        test_generar_excel_transversal,
        test_roundtrip_misional,
        test_roundtrip_territorial,
        test_deteccion_familia,
        test_consolidar,
        test_lectura_real_hub,
        test_cero_iterrows,
        test_cero_iterrows_constructores,
    ]

    for test_fn in tests:
        nombre = test_fn.__name__
        try:
            print(f"  🔍 {nombre}...")
            test_fn()
        except Exception as e:
            global _FAILED
            _FAILED += 1
            _ERRORS.append(f"  💥 {nombre} — EXCEPCIÓN: {e}")
            print(f"  💥 {nombre} — EXCEPCIÓN: {e}")

    return _resumen()


if __name__ == '__main__':
    success = run_all()
    sys.exit(0 if success else 1)
